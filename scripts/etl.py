"""
Seed the Wharton Alum Champions League app from the commissioner's Excel workbook.

Usage:
    python scripts/etl.py "path/to/FFL Stats and Keepers_2026 PreSeason.xlsx"

Writes JSON into public/data/. Re-runnable: files that the app owns at runtime
(cash.json, faab.json, trade-queue.json) are never overwritten if they exist.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "public" / "data"

# Managers are recorded by surname in the stats sheets and by first name in the
# trade sheets. This mapping is derived from the 2025 standings sheet (team ->
# surname) joined to Keepers '26 (team -> first name).
MANAGERS = [
    # surname,      first,       2025 team,               active
    ("Varghese",    "Tajar",     "Malikety Split",        True),
    ("Waldman",     "Stu",       "I Must Break You",      True),
    ("Tollinche",   "Felix",     "El Morro",              True),
    ("Buzik",       "Alex",      "SuperBowlJesus",        True),
    ("Bernstein",   "Bernstein", "Orchids of Asia",       True),
    ("Baugh",       "Baugh",     "Juggernaut",            True),
    ("Velamoor",    "Velamoor",  "Private Sector",        True),
    ("Mukheja",     "Anik",      "Karma's L-Ah-Ma-tize",  True),
    ("Snyder",      "Nate",      "Whobody Wants It?",     True),
    ("Incognito",   "Jim",       "no ragrets",            True),
    ("Konciak",     "Dave",      "Bridge N Tunnel",       True),
    ("Lalwani",     "Lalwani",   "Quien Es Tu Papa",      True),
    ("Evans",       "Evans",     None,                    False),
    ("Banerjee",    "Banerjee",  None,                    False),
    ("Kim",         "Kim",       None,                    False),
    ("Kurucz",      "Kurucz",    None,                    False),
]

FIRST_TO_SURNAME = {first: last for last, first, _, _ in MANAGERS}
SEASON_YEARS = list(range(2004, 2026))
KEEPER_YEARS = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]

BASE_DRAFT_BUDGET = 200
BASE_FAAB_BUDGET = 100


def slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def clean(value):
    """Normalise a cell into str | float | None."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("�", "'").replace(" ", " ").strip()
        if text in ("", "NA", "N/A", "-", "—"):
            return None
        return text
    return value


def num(value):
    """Numbers in this workbook are inconsistently stored as text."""
    value = clean(value)
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    if isinstance(value, str):
        stripped = value.replace("$", "").replace(",", "").strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", stripped):
            return round(float(stripped), 4)
    return None


def as_year(value):
    parsed = num(value)
    if parsed is not None and 2000 < parsed < 2100 and float(parsed).is_integer():
        return int(parsed)
    return None


def build_managers():
    return [
        {
            "id": slug(last),
            "surname": last,
            "firstName": first,
            "displayName": first if first != last else last,
            "team": team,
            "active": active,
        }
        for last, first, team, active in MANAGERS
    ]


def build_seasons(wb):
    """One record per team-season from the per-year sheets."""
    seasons = []
    for year in SEASON_YEARS:
        sheet = wb[str(year)]
        teams = []
        for row in sheet.iter_rows(min_row=7, max_row=40, max_col=12, values_only=True):
            rank, team, manager = num(row[0]), clean(row[1]), clean(row[2])
            if rank is None or not isinstance(manager, str):
                continue
            if manager not in FIRST_TO_SURNAME.values():
                continue
            teams.append(
                {
                    "rank": int(rank),
                    "team": team,
                    "manager": slug(manager),
                    "wins": num(row[3]) or 0,
                    "losses": num(row[4]) or 0,
                    "pointsFor": num(row[6]),
                    "pointsAgainst": num(row[7]),
                    "avgPointsFor": num(row[8]),
                    "avgPointsAgainst": num(row[9]),
                    "playoffWins": num(row[10]) or 0,
                    "playoffLosses": num(row[11]) or 0,
                }
            )
        if not teams:
            continue
        teams.sort(key=lambda t: t["rank"])
        seasons.append(
            {
                "year": year,
                "teamCount": len(teams),
                "keeperEra": year >= 2014,
                "champion": teams[0]["manager"],
                "runnerUp": teams[1]["manager"] if len(teams) > 1 else None,
                "thirdPlace": teams[2]["manager"] if len(teams) > 2 else None,
                "teams": teams,
            }
        )
    seasons.sort(key=lambda s: -s["year"])
    return seasons


def _keeper_col(header_row):
    """Keeper block columns shift between workbook vintages; find 'Player'."""
    for idx, cell in enumerate(header_row):
        if idx >= 3 and clean(cell) == "Player":
            return idx
    return None


def build_keepers(wb):
    """Ending rosters + selected keepers + budget math, per team per year."""
    by_year = {}
    for year in KEEPER_YEARS:
        sheet = wb[f"Keepers '{str(year)[2:]}"]
        grid = [list(r) for r in sheet.iter_rows(max_col=20, values_only=True)]
        blocks = []
        row_index = 0
        while row_index < len(grid):
            row = grid[row_index]
            if clean(row[1]) == "Keeper Cost" and clean(row[0]):
                team = clean(row[0])
                kcol = _keeper_col(row)
                block = {
                    "team": team,
                    "manager": None,
                    "endingRoster": [],
                    "keepers": [],
                    "keeperSalary": None,
                    "cashTraded": 0,
                    "draftBudget": None,
                }
                cursor = row_index + 1
                while cursor < len(grid):
                    line = grid[cursor]
                    is_new_block = clean(line[1]) == "Keeper Cost" and clean(line[0])
                    if is_new_block:
                        break
                    player = clean(line[0])
                    if player:
                        block["endingRoster"].append(
                            {
                                "player": player,
                                "cost": num(line[1]),
                                "contractYear": clean(line[2]),
                            }
                        )
                    if kcol is not None and kcol < len(line):
                        label = clean(line[kcol])
                        value = num(line[kcol + 1]) if kcol + 1 < len(line) else None
                        if label == "Keeper Salary":
                            block["keeperSalary"] = value
                            # The manager's first name sits just left of the label.
                            for back in range(1, 4):
                                candidate = clean(line[kcol - back])
                                if isinstance(candidate, str) and candidate in FIRST_TO_SURNAME:
                                    block["manager"] = slug(FIRST_TO_SURNAME[candidate])
                                    break
                        elif label == "Cash Traded":
                            block["cashTraded"] = value or 0
                            for back in range(1, 4):
                                candidate = clean(line[kcol - back])
                                if isinstance(candidate, str) and candidate in FIRST_TO_SURNAME:
                                    block["manager"] = slug(FIRST_TO_SURNAME[candidate])
                                    break
                        elif label == "Draft Budget":
                            block["draftBudget"] = value
                        elif label and label != "Player":
                            block["keepers"].append(
                                {
                                    "player": label,
                                    "salary": value,
                                    "contractYear": clean(line[kcol + 2])
                                    if kcol + 2 < len(line)
                                    else None,
                                }
                            )
                    cursor += 1
                blocks.append(block)
                row_index = cursor
            else:
                row_index += 1
        if blocks:
            by_year[str(year)] = blocks
    return by_year


def backfill_keeper_managers(keepers, seasons):
    """
    Only the most recent keeper sheet names the manager beside the budget math.
    For every other year, the block's team is the roster as it stood at the end
    of the prior season, so the standings sheet for year-1 supplies the owner.
    """
    by_year = {}
    for season in seasons:
        for team in season["teams"]:
            if team["team"]:
                by_year.setdefault(season["year"], {})[team["team"].strip().lower()] = team[
                    "manager"
                ]

    # A team name used by exactly one manager across all 22 seasons is a safe
    # fallback when that season's own standings sheet used a different name.
    global_names = {}
    for season in seasons:
        for team in season["teams"]:
            if team["team"]:
                global_names.setdefault(team["team"].strip().lower(), set()).add(team["manager"])
    unique_names = {
        name: next(iter(owners)) for name, owners in global_names.items() if len(owners) == 1
    }

    resolved = 0
    for year_key, blocks in keepers.items():
        lookup = by_year.get(int(year_key) - 1, {})
        for block in blocks:
            if block["manager"] or not block["team"]:
                continue
            name = block["team"].strip().lower()
            manager = lookup.get(name)
            if not manager:
                # Team names drift between sheets; fall back to a prefix match.
                for known, candidate in lookup.items():
                    if known[:10] == name[:10]:
                        manager = candidate
                        break
            if not manager:
                manager = unique_names.get(name)
            if manager:
                block["manager"] = manager
                resolved += 1
    return resolved


def build_waivers(wb, team_to_manager):
    """FAAB waiver claims, where the workbook recorded them (2025 onward)."""

    def resolve(team):
        if not team:
            return None
        if team in team_to_manager:
            return team_to_manager[team]
        # Team names drift slightly between sheets (truncation, stray encoding).
        for known, manager in team_to_manager.items():
            if known[:12].lower() == team[:12].lower():
                return manager
        return None

    claims = []
    for year in KEEPER_YEARS:
        sheet = wb[f"Keepers '{str(year)[2:]}"]
        grid = [list(r) for r in sheet.iter_rows(max_col=20, values_only=True)]
        header = None
        for row in grid[:6]:
            for idx, cell in enumerate(row):
                if clean(cell) == "Waivers":
                    header = idx
        if header is None:
            continue
        # The waiver column on a Keepers sheet is filled in during that same
        # season (Keepers '25 carries the 2025 season's FAAB claims).
        season = year
        for row in grid:
            player = clean(row[header]) if header < len(row) else None
            bid = num(row[header + 1]) if header + 1 < len(row) else None
            team = clean(row[header + 2]) if header + 2 < len(row) else None
            if not player or player == "Waivers" or bid is None:
                continue
            claims.append(
                {
                    "season": season,
                    "player": player,
                    "bid": bid,
                    "team": team,
                    "manager": resolve(team),
                }
            )
    return claims


def build_trades(wb):
    """Structured trades with multi-year auction-dollar obligations."""
    sheet = wb["Trade Detail"]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    header = rows[0]
    year_cols = {}
    for idx, cell in enumerate(header):
        year = as_year(cell)
        if year:
            year_cols[idx] = year

    trades = []
    for i, row in enumerate(rows[1:], start=2):
        batch, seller, buyer = clean(row[0]), clean(row[1]), clean(row[2])
        players = clean(row[3])
        if not batch or not seller or seller == "LEAGUE TOTAL" or not buyer:
            continue
        obligations = []
        for idx, year in year_cols.items():
            amount = num(row[idx]) if idx < len(row) else None
            if amount:
                obligations.append({"year": year, "amount": amount})
        season = int(re.match(r"(\d{4})", batch).group(1))
        trades.append(
            {
                "id": f"t{season}-{i:03d}",
                "batch": batch,
                "season": season,
                "preseason": "Preseason" in batch,
                "seller": slug(FIRST_TO_SURNAME.get(seller, seller)),
                "buyer": slug(FIRST_TO_SURNAME.get(buyer, buyer)),
                "players": players,
                "terms": clean(row[4]),
                "totalDollars": num(row[5]) or 0,
                "obligations": obligations,
                "status": "approved",
                "source": "workbook",
            }
        )
    return trades


def build_legacy_trades(wb):
    """Free-text trade log, 2013-2022, grouped by its section headings."""
    sheet = wb["Trades"]
    groups = []
    current = None
    for (cell,) in sheet.iter_rows(max_col=1, values_only=True):
        text = clean(cell)
        if not text or text == "Wharton Alum Champions League":
            continue
        if re.match(r"^\d{4} Trades", text):
            current = {"heading": text, "entries": []}
            groups.append(current)
        elif current:
            current["entries"].append(text)
    return groups


def build_rules(wb):
    sheet = wb["Rules"]
    lines = [clean(c) for (c,) in sheet.iter_rows(max_col=1, values_only=True)]
    return [line for line in lines if line]


def build_trade_ledger(trades):
    """Per-manager auction-dollar received/sent/net by obligation year."""
    ledger = {}
    for trade in trades:
        for ob in trade["obligations"]:
            year = str(ob["year"])
            ledger.setdefault(year, {})
            for manager, key in ((trade["seller"], "received"), (trade["buyer"], "sent")):
                entry = ledger[year].setdefault(
                    manager, {"received": 0.0, "sent": 0.0, "net": 0.0}
                )
                entry[key] += ob["amount"]
    for year in ledger:
        for entry in ledger[year].values():
            entry["net"] = round(entry["received"] - entry["sent"], 2)
            entry["received"] = round(entry["received"], 2)
            entry["sent"] = round(entry["sent"], 2)
    return ledger


def verify_ledger(wb, ledger):
    """Cross-check the derived ledger against the workbook's Trade $$$ tab."""
    sheet = wb["Trade $$$"]
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    years = {}
    for idx, cell in enumerate(rows[1]):
        year = as_year(cell)
        if year:
            years[idx] = year
    problems = []
    for row in rows[3:]:
        manager = clean(row[0])
        if not manager or manager == "LEAGUE TOTAL":
            continue
        mid = slug(FIRST_TO_SURNAME.get(manager, manager))
        for idx, year in years.items():
            expected_net = num(row[idx + 2]) or 0
            actual = ledger.get(str(year), {}).get(mid, {}).get("net", 0)
            if abs(expected_net - actual) > 0.01:
                problems.append(f"{manager} {year}: workbook {expected_net} vs derived {actual}")
    return problems


def write(name, payload, overwrite=True):
    path = OUT / name
    if not overwrite and path.exists():
        print(f"  skip  {name} (app-owned, already exists)")
        return
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    size = path.stat().st_size
    print(f"  write {name} ({size:,} bytes)")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    source = Path(sys.argv[1])
    wb = openpyxl.load_workbook(source, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)

    managers = build_managers()
    seasons = build_seasons(wb)
    keepers = build_keepers(wb)
    trades = build_trades(wb)
    ledger = build_trade_ledger(trades)

    resolved = backfill_keeper_managers(keepers, seasons)
    unresolved = [
        f"{year}:{block['team']}"
        for year, blocks in keepers.items()
        for block in blocks
        if not block["manager"]
    ]
    print(f"Backfilled manager on {resolved} keeper blocks from the standings sheets.")
    if unresolved:
        print(f"!! {len(unresolved)} keeper blocks still unattributed: {', '.join(unresolved[:8])}")

    problems = verify_ledger(wb, ledger)
    if problems:
        print("\n!! derived trade ledger disagrees with the workbook:")
        for problem in problems:
            print("   ", problem)
    else:
        print("\nOK derived trade ledger matches the workbook's Trade $$$ tab exactly")

    print(f"\nWriting to {OUT}")
    write("managers.json", managers)
    write("seasons.json", seasons)
    write("keepers.json", keepers)
    write("trades.json", trades)
    write("trade-ledger.json", ledger)
    team_to_manager = {
        block["team"]: block["manager"]
        for blocks in keepers.values()
        for block in blocks
        if block["team"] and block["manager"]
    }
    write("waivers.json", build_waivers(wb, team_to_manager))
    write("legacy-trades.json", build_legacy_trades(wb))
    write("rules.json", build_rules(wb))
    write(
        "league.json",
        {
            "name": "Wharton Alum Champions League",
            "commissioner": "Tajar Varghese",
            "currentSeason": 2026,
            "baseDraftBudget": BASE_DRAFT_BUDGET,
            "baseFaabBudget": BASE_FAAB_BUDGET,
            "keeperSlots": 4,
            "maxContractYears": 4,
            "faabScale": [
                {"maxPct": 20, "keeperCost": 5},
                {"maxPct": 40, "keeperCost": 10},
                {"maxPct": 60, "keeperCost": 15},
                {"maxPct": 100, "keeperCost": 20},
            ],
            "sourceWorkbook": source.name,
        },
    )

    # App-owned files: seeded once, then maintained through the UI.
    write("cash.json", {"entries": []}, overwrite=False)
    write("faab.json", {"season": 2026, "entries": []}, overwrite=False)
    write("trade-queue.json", {"proposals": []}, overwrite=False)

    print(
        f"\nSeeded {len(managers)} managers, {len(seasons)} seasons, "
        f"{sum(len(v) for v in keepers.values())} keeper blocks, {len(trades)} trades."
    )


if __name__ == "__main__":
    main()
